#!/usr/bin/env python3
"""
Execute:
python bulk_from_excel_map_junction_insertions.py \
  --xlsx /Users/ecreed/Desktop/KelloggRotation/CodeProjects/TagmentationAnalysis/01302026_bulk/01302026W1DN1/W1DN1_R1__bulk_inputs.xlsx \
  --sheet Sheet1 \
  --report-dir /Users/ecreed/Desktop/KelloggRotation/CodeProjects/TagmentationAnalysis/01302026_bulk/01302026W1DN1/W1DN1_Output


Bulk runner for 01302026_filtered_map_junction_insertions.py-style pipeline.

Reads an Excel sheet with per-sample parameters and runs:
  0) QC filter reads by Phred scores (FASTQ-level)
  1) cutadapt trim donor
  2) bwa mem map + sort/index
  3) call insertion per read
  4) site-level min-read-count filter -> writes filtered_<FASTQ_STEM>.tsv

Sample naming:
  samplename is ALWAYS derived from the FASTQ filename stem (Path(fastq).stem).
  No Excel samplename column is used or supported.

Outputs (written to --report-dir):
  - filtered_tsv_manifest.xlsx : one row per successfully produced filtered TSV
  - removed_samples.xlsx       : samples removed (QC empty, min-read-count empty, or errors)

Notes:
- Core analysis functions are copied from your single-sample script without changing logic.
- File-level QC removal criterion implemented as: QC-passing FASTQ has 0 reads -> removed.
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter
import gzip
import subprocess
from pathlib import Path
from typing import Dict, Optional, Tuple, Any

import pysam
from openpyxl import load_workbook, Workbook


# ----------------------------
# Unchanged helpers / pipeline
# ----------------------------

def run(cmd, *, shell=False):
    cmd_str = cmd if isinstance(cmd, str) else " ".join(map(str, cmd))
    print(f"[cmd] {cmd_str}")
    subprocess.run(cmd, check=True, shell=shell)


def open_maybe_gz(path: Path, mode: str):
    if str(path).endswith(".gz"):
        return gzip.open(path, mode)
    return open(path, mode)


def phred_scores(qual: str) -> list[int]:
    return [ord(c) - 33 for c in qual.rstrip("\n")]


def ensure_bwa_index(ref_fa: Path):
    needed = [ref_fa.with_suffix(ref_fa.suffix + ext) for ext in [".amb", ".ann", ".bwt", ".pac", ".sa"]]
    if all(p.exists() for p in needed):
        return
    run(["bwa", "index", str(ref_fa)])


def qc_filter_fastq(
    fastq_in: Path,
    fastq_out: Path,
    *,
    lowq_thresh: int = 20,
    max_lowq_frac: float = 0.50,
) -> Dict[str, Tuple[int, float, float, float]]:
    """
    QC gate BEFORE mapping:
      drop if >= max_lowq_frac of bases have Q < lowq_thresh

    Returns dict:
      read_id -> (read_len, mean_q, frac_q_lt20, pct_q_ge30)
    """
    metrics: Dict[str, Tuple[int, float, float, float]] = {}

    n_total = 0
    n_pass = 0

    with open_maybe_gz(fastq_in, "rt") as fin, open_maybe_gz(fastq_out, "wt") as fout:
        while True:
            h = fin.readline()
            if not h:
                break
            s = fin.readline()
            p = fin.readline()
            q = fin.readline()
            if not (s and p and q):
                raise ValueError("Truncated FASTQ record detected.")

            n_total += 1
            rid = h.strip().split()[0]
            if rid.startswith("@"):
                rid = rid[1:]

            qs = phred_scores(q)
            if not qs:
                continue

            read_len = len(qs)
            mean_q = sum(qs) / read_len
            n_lt = sum(1 for x in qs if x < lowq_thresh)
            frac_lt = n_lt / read_len
            n_ge30 = sum(1 for x in qs if x >= 30)
            pct_ge30 = 100.0 * (n_ge30 / read_len)

            metrics[rid] = (read_len, mean_q, frac_lt, pct_ge30)

            if frac_lt >= max_lowq_frac:
                continue  # drop

            fout.write(h)
            fout.write(s)
            fout.write(p)
            fout.write(q)
            n_pass += 1

    print(f"[info] QC scanned reads: {n_total}")
    print(f"[info] QC passed reads: {n_pass}")
    if n_total:
        print(f"[info] QC pass rate: {n_pass / n_total:.4f}")
    print(f"[info] QC FASTQ: {fastq_out}")
    return metrics


def cutadapt_trim(
    fastq_in: Path,
    fastq_out: Path,
    donor_seq: str,
    donor_side: str,
    min_len: int,
    error_rate: float,
    cores: int,
):
    if donor_side == "5p":
        adapter_arg = ["-g", donor_seq]
    else:
        adapter_arg = ["-a", donor_seq]

    cmd = [
        "cutadapt",
        *adapter_arg,
        "--discard-untrimmed",
        "-e", str(error_rate),
        "-m", str(min_len),
        "-j", str(cores),
        "-o", str(fastq_out),
        str(fastq_in),
    ]
    run(cmd)


def bwa_map(ref_fa: Path, fastq: Path, bam_out: Path, cores: int):
    sam_tmp = bam_out.with_suffix(".sam")
    run(f"bwa mem -t {cores} {ref_fa} {fastq} > {sam_tmp}", shell=True)
    run(["samtools", "sort", "-@", str(cores), "-o", str(bam_out), str(sam_tmp)])
    run(["samtools", "index", str(bam_out)])
    try:
        sam_tmp.unlink()
    except FileNotFoundError:
        pass


def ref_aln_span(read: pysam.AlignedSegment) -> tuple[int, int]:
    start0 = read.reference_start
    end0 = read.reference_end
    return start0, end0


def insertion_coord(read: pysam.AlignedSegment, donor_side: str) -> Optional[int]:
    if read.is_unmapped:
        return None
    if donor_side == "5p":
        return read.reference_start
    else:
        if read.reference_end is None:
            return None
        return read.reference_end - 1


def is_perfect_genome_match_allow_clips(read: pysam.AlignedSegment) -> bool:
    """
    Enforce NM==0; allow clipping (S/H) as deviations.
    """
    nm = read.get_tag("NM") if read.has_tag("NM") else None
    if nm is None or nm != 0:
        return False
    return True


def call_insertions(
    *,
    bam_path: Path,
    out_tsv: Path,
    donor_side: str,
    min_mapq: int,
    primary_only: bool,
    qc_metrics: Dict[str, Tuple[int, float, float, float]],
):
    header = [
        "read_id", "ref", "ins0", "ins1", "strand", "mapq", "cigar", "aln_start0", "aln_end0_excl",
        "qc_read_len", "qc_mean_q", "qc_frac_q_lt20", "qc_pct_q_ge30",
    ]

    bam = pysam.AlignmentFile(str(bam_path), "rb")
    n_total = 0
    n_kept = 0
    n_not_perfect = 0
    n_no_qc = 0

    with open(out_tsv, "w", newline="") as f:
        f.write("\t".join(header) + "\n")

        for read in bam.fetch(until_eof=True):
            n_total += 1

            if read.is_unmapped:
                continue
            if read.mapping_quality < min_mapq:
                continue
            if primary_only and (read.is_secondary or read.is_supplementary):
                continue

            if not is_perfect_genome_match_allow_clips(read):
                n_not_perfect += 1
                continue

            ins0 = insertion_coord(read, donor_side=donor_side)
            if ins0 is None:
                continue

            ref = bam.get_reference_name(read.reference_id)
            strand = "-" if read.is_reverse else "+"
            start0, end0_excl = ref_aln_span(read)

            rid = read.query_name
            qc = qc_metrics.get(rid)
            if qc is None:
                n_no_qc += 1
                qc_read_len, qc_mean_q, qc_frac_lt, qc_pct_ge30 = ("NA", "NA", "NA", "NA")
            else:
                qc_read_len, qc_mean_q, qc_frac_lt, qc_pct_ge30 = qc
                qc_mean_q = f"{qc_mean_q:.3f}"
                qc_frac_lt = f"{qc_frac_lt:.4f}"
                qc_pct_ge30 = f"{qc_pct_ge30:.2f}"

            f.write("\t".join(map(str, [
                rid, ref, ins0, ins0 + 1, strand, read.mapping_quality,
                read.cigarstring, start0, end0_excl,
                qc_read_len, qc_mean_q, qc_frac_lt, qc_pct_ge30,
            ])) + "\n")
            n_kept += 1

    bam.close()
    print(f"[info] BAM reads scanned: {n_total}")
    print(f"[info] Insertions written: {n_kept}")
    print(f"[info] Dropped (not perfect genome match): {n_not_perfect}")
    print(f"[info] Missing QC metrics in dict: {n_no_qc}")
    print(f"[info] Output TSV: {out_tsv}")


def filter_min_read_count_sites(
    tsv_in: Path,
    tsv_out: Path,
    *,
    min_read_count: int,
) -> tuple[int, int, int]:
    if min_read_count <= 0:
        raise ValueError("min_read_count must be > 0 to use this filter")

    site_counts: Counter[tuple[str, int]] = Counter()
    n_in = 0

    with open(tsv_in, "r", newline="") as fin:
        reader = csv.DictReader(fin, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"Empty TSV or missing header: {tsv_in}")
        if "ref" not in reader.fieldnames or "ins0" not in reader.fieldnames:
            raise ValueError(f"TSV missing required columns 'ref' and/or 'ins0': {tsv_in}")

        for row in reader:
            n_in += 1
            ref = row.get("ref")
            ins0_s = row.get("ins0")
            if not ref or ins0_s in (None, ""):
                continue
            try:
                ins0 = int(ins0_s)
            except ValueError:
                continue
            site_counts[(ref, ins0)] += 1

    kept_sites = {k for k, c in site_counts.items() if c >= min_read_count}

    n_out = 0
    with open(tsv_in, "r", newline="") as fin, open(tsv_out, "w", newline="") as fout:
        reader = csv.DictReader(fin, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"Empty TSV or missing header: {tsv_in}")
        writer = csv.DictWriter(fout, fieldnames=reader.fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()

        for row in reader:
            ref = row.get("ref")
            ins0_s = row.get("ins0")
            if not ref or ins0_s in (None, ""):
                continue
            try:
                ins0 = int(ins0_s)
            except ValueError:
                continue
            if (ref, ins0) in kept_sites:
                writer.writerow(row)
                n_out += 1

    return n_in, n_out, len(kept_sites)


# ----------------------------
# Bulk / Excel I/O layer
# ----------------------------

REQUIRED_COLS = [
    "fastq", "ref", "outdir", "donor-seq", "donor-side",
    "min-read-count", "min-mapq", "qc-lowq-thresh", "qc-max-lowq-frac",
]


def count_fastq_records(fq: Path) -> int:
    """Counts FASTQ records by counting lines / 4 (supports .gz)."""
    n_lines = 0
    with open_maybe_gz(fq, "rt") as f:
        for _ in f:
            n_lines += 1
    return n_lines // 4


def read_excel_rows(xlsx: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    col_idx = {name: i for i, name in enumerate(header) if name}

    missing = [c for c in REQUIRED_COLS if c not in col_idx]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue

        d: dict[str, Any] = {}
        for c in REQUIRED_COLS:
            d[c] = r[col_idx[c]]
        out.append(d)

    return out


def write_xlsx(path: Path, header: list[str], data_rows: list[list[Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def norm_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def norm_int(x: Any) -> int:
    if x is None or str(x).strip() == "":
        raise ValueError("Expected integer but got blank.")
    return int(x)


def norm_float(x: Any) -> float:
    if x is None or str(x).strip() == "":
        raise ValueError("Expected float but got blank.")
    return float(x)


def process_one_sample(row: dict[str, Any]) -> tuple[bool, dict[str, Any]]:
    """
    Returns (ok, info)
    ok=True  -> info contains manifest fields including filtered_tsv path
    ok=False -> info contains removal reason
    """
    fastq = Path(norm_str(row["fastq"]))
    ref = Path(norm_str(row["ref"]))
    outdir = Path(norm_str(row["outdir"]))
    donor_seq = norm_str(row["donor-seq"])
    donor_side = norm_str(row["donor-side"])
    min_read_count = norm_int(row["min-read-count"])
    min_mapq = norm_int(row["min-mapq"])
    qc_lowq_thresh = norm_int(row["qc-lowq-thresh"])
    qc_max_lowq_frac = norm_float(row["qc-max-lowq-frac"])

    # Sample name ALWAYS from FASTQ filename
    samplename = fastq.stem

    outdir.mkdir(parents=True, exist_ok=True)

    try:
        ensure_bwa_index(ref)

        # 0) QC filter (read-level)
        qc_fastq = outdir / f"{fastq.stem}.qc.fastq"
        qc_metrics = qc_filter_fastq(
            fastq,
            qc_fastq,
            lowq_thresh=qc_lowq_thresh,
            max_lowq_frac=qc_max_lowq_frac,
        )

        # File-level QC removal: if QC output has 0 reads, remove this sample
        qc_pass_reads = count_fastq_records(qc_fastq)
        if qc_pass_reads == 0:
            return False, {
                "samplename": samplename,
                "fastq": str(fastq),
                "reason": "qc_failed_no_passing_reads",
                "detail": f"QC output FASTQ contained 0 reads (lowq_thresh={qc_lowq_thresh}, max_lowq_frac={qc_max_lowq_frac})",
                "outdir": str(outdir),
            }

        # 1) Trim donor
        trimmed_fastq = outdir / f"{fastq.stem}.trimmed.fastq"
        cutadapt_trim(
            qc_fastq,
            trimmed_fastq,
            donor_seq=donor_seq,
            donor_side=donor_side,
            min_len=20,
            error_rate=0.1,
            cores=8,
        )

        # 2-3) Map + sort/index
        bam = outdir / f"{fastq.stem}.sorted.bam"
        bwa_map(ref, trimmed_fastq, bam, cores=8)

        # 4) Call insertions per read
        raw_tsv = outdir / f"{fastq.stem}.insertions.{donor_side}.tsv"
        call_insertions(
            bam_path=bam,
            out_tsv=raw_tsv,
            donor_side=donor_side,
            min_mapq=min_mapq,
            primary_only=False,
            qc_metrics=qc_metrics,
        )

        # 5) Min-read-count site-level filter -> filtered_<FASTQ_STEM>.tsv
        filtered_tsv = outdir / f"filtered_{samplename}.tsv"
        n_in, n_out, n_sites = filter_min_read_count_sites(
            raw_tsv, filtered_tsv, min_read_count=min_read_count
        )

        if n_out == 0:
            # Remove from analysis; track in removed sheet
            try:
                filtered_tsv.unlink(missing_ok=True)
            except TypeError:
                if filtered_tsv.exists():
                    filtered_tsv.unlink()

            return False, {
                "samplename": samplename,
                "fastq": str(fastq),
                "reason": "min_read_count_removed_no_sites",
                "detail": f"min_read_count={min_read_count} yielded 0 output rows (input_rows={n_in}, sites_kept={n_sites})",
                "outdir": str(outdir),
            }

        return True, {
            "samplename": samplename,
            "filtered_tsv": str(filtered_tsv),
            "fastq": str(fastq),
            "ref": str(ref),
            "outdir": str(outdir),
            "donor_seq": donor_seq,
            "donor_side": donor_side,
            "min_read_count": min_read_count,
            "min_mapq": min_mapq,
            "qc_lowq_thresh": qc_lowq_thresh,
            "qc_max_lowq_frac": qc_max_lowq_frac,
            "qc_pass_reads": qc_pass_reads,
            "raw_rows_in": n_in,
            "filtered_rows_out": n_out,
            "sites_kept": n_sites,
        }

    except Exception as e:
        return False, {
            "samplename": samplename,
            "fastq": str(fastq),
            "reason": "runtime_error",
            "detail": repr(e),
            "outdir": str(outdir),
        }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path, help="Excel file with batch inputs.")
    ap.add_argument("--sheet", required=True, help="Sheet name inside the Excel workbook.")
    ap.add_argument("--report-dir", required=True, type=Path, help="Directory for summary Excel outputs.")
    args = ap.parse_args()

    rows = read_excel_rows(args.xlsx, args.sheet)

    kept_manifest: list[dict[str, Any]] = []
    removed: list[dict[str, Any]] = []

    for i, row in enumerate(rows, start=1):
        print(f"\n=== Sample {i}/{len(rows)} ===")
        ok, info = process_one_sample(row)
        if ok:
            kept_manifest.append(info)
        else:
            removed.append(info)

    # Write manifest Excel (filtered TSVs)
    manifest_path = args.report_dir / "filtered_tsv_manifest.xlsx"
    manifest_header = [
        "samplename", "filtered_tsv", "fastq", "ref", "outdir",
        "donor_seq", "donor_side", "min_read_count", "min_mapq",
        "qc_lowq_thresh", "qc_max_lowq_frac",
        "qc_pass_reads", "raw_rows_in", "filtered_rows_out", "sites_kept",
    ]
    manifest_rows = [[m.get(h, "") for h in manifest_header] for m in kept_manifest]
    write_xlsx(manifest_path, manifest_header, manifest_rows)
    print(f"[info] Wrote: {manifest_path}")

    # Write removed Excel
    removed_path = args.report_dir / "removed_samples.xlsx"
    removed_header = ["samplename", "fastq", "outdir", "reason", "detail"]
    removed_rows = [[r.get(h, "") for h in removed_header] for r in removed]
    write_xlsx(removed_path, removed_header, removed_rows)
    print(f"[info] Wrote: {removed_path}")

    print(f"\n[summary] kept={len(kept_manifest)} removed={len(removed)} total={len(rows)}")


if __name__ == "__main__":
    main()
